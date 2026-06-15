"""
CAMT.053 -> Excel (.xlsx) exporter.

Renders the parsed statements from camt_reader.parse_camt053 into a clean,
on-brand workbook (one sheet per statement): brand-orange header bar, bold
title + italic grey subtitle, a small balance-reconciliation block built with
LIVE formulas (opening + SUM(entries) = expected closing), then a frozen,
auto-filterable entries table in Verdana 11.

No third-party dependencies beyond openpyxl (already a converter dependency).
"""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ORANGE = 'EB5F36'
GREY = '8A8A8A'
LINE = 'E6E9EF'

_FONT = 'Verdana'
_title_font = Font(name=_FONT, size=14, bold=True, color='FFFFFF')
_sub_font = Font(name=_FONT, size=9, italic=True, color=GREY)
_hdr_font = Font(name=_FONT, size=11, bold=True, color='FFFFFF')
_lbl_font = Font(name=_FONT, size=11, bold=True)
_cell_font = Font(name=_FONT, size=11)
_orange_fill = PatternFill('solid', fgColor=ORANGE)
_hdr_fill = PatternFill('solid', fgColor=ORANGE)
_thin_bottom = Border(bottom=Side(style='thin', color=LINE))
_money_fmt = '#,##0.00'

_COLS = ['Booking date', 'Value date', 'Direction', 'Amount', 'Currency', 'Description', 'Reference']
_WIDTHS = [14, 14, 11, 15, 9, 60, 22]


def _amount(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def build_xlsx(statements, source_name='statement'):
    """statements: list of dicts from camt_reader. Returns xlsx bytes."""
    today = date.today().isoformat()
    wb = Workbook()
    wb.remove(wb.active)

    for i, s in enumerate(statements):
        title = (s.get('account') or 'Account')[:28]
        ws = wb.create_sheet(title=(f'Statement {i + 1}' if len(statements) > 1 else title) or 'Statement')
        _render_sheet(ws, s, source_name, today)

    wb.properties.creator = 'Daniel Buetler'
    wb.properties.lastModifiedBy = 'Daniel Buetler'

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_sheet(ws, s, source_name, today):
    ncols = len(_COLS)
    last_col = get_column_letter(ncols)

    # Title bar (row 1) — brand orange, white bold, merged across the table.
    ws.merge_cells(f'A1:{last_col}1')
    c = ws['A1']
    c.value = f"CAMT.053 Statement — {s.get('account') or '—'}"
    c.font = _title_font
    c.fill = _orange_fill
    c.alignment = Alignment(vertical='center', horizontal='left', indent=1)
    ws.row_dimensions[1].height = 26
    for col in range(1, ncols + 1):
        ws.cell(row=1, column=col).fill = _orange_fill

    # Subtitle (row 2) — italic grey: owner · ccy · period · count · generated.
    ws.merge_cells(f'A2:{last_col}2')
    bits = [b for b in [
        s.get('owner'),
        s.get('currency'),
        f"{s.get('period_from') or '?'} → {s.get('period_to') or '?'}",
        f"{len(s.get('entries') or [])} entries",
        f"generated {today}",
    ] if b]
    sc = ws['A2']
    sc.value = '   ·   '.join(bits)
    sc.font = _sub_font
    sc.alignment = Alignment(vertical='center', horizontal='left', indent=1)
    ws.row_dimensions[2].height = 16

    # Balance reconciliation block (rows 4-8), built with LIVE formulas.
    # The entries table starts at header_row; amounts live in column D.
    header_row = 10
    first_data = header_row + 1
    n = len(s.get('entries') or [])
    last_data = first_data + n - 1
    amt_range = f'D{first_data}:D{last_data}' if n else None
    opening = _amount(s.get('opening')) if s.get('opening') is not None else None
    closing = _amount(s.get('closing')) if s.get('closing') is not None else None

    recon_rows = [
        ('Opening balance (OPBD)', opening if opening is not None else 0.0, False),
        ('Movements (Σ entries)', f'=SUM({amt_range})' if amt_range else 0.0, True),
        ('Expected closing', '=B4+B5', True),
        ('Stated closing (CLBD)', closing if closing is not None else 0.0, False),
        ('Difference (stated − expected)', '=B7-B6', True),
    ]
    for idx, (label, value, is_formula) in enumerate(recon_rows):
        r = 4 + idx
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = _lbl_font
        vc = ws.cell(row=r, column=2, value=value)
        vc.font = _cell_font
        vc.number_format = _money_fmt

    # Table header (frozen) + auto-filter.
    for col, name in enumerate(_COLS, start=1):
        hc = ws.cell(row=header_row, column=col, value=name)
        hc.font = _hdr_font
        hc.fill = _hdr_fill
        hc.alignment = Alignment(vertical='center', horizontal='left', indent=1)
    ws.row_dimensions[header_row].height = 20

    for ridx, e in enumerate(s.get('entries') or []):
        r = first_data + ridx
        row = [
            e.get('booking_date') or '',
            e.get('value_date') or '',
            e.get('cd_ind') or '',
            _amount(e.get('amount')),
            e.get('currency') or s.get('currency') or '',
            e.get('description') or '',
            e.get('ref') or '',
        ]
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = _cell_font
            cell.border = _thin_bottom
            if col == 4:
                cell.number_format = _money_fmt

    # Column widths, freeze below the header, filter over the table.
    for col, w in enumerate(_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f'A{first_data}'
    end = last_data if n else header_row
    ws.auto_filter.ref = f'A{header_row}:{last_col}{end}'
